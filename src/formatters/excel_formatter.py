# src/formatters/excel_formatter.py
# Excel 格式化器 (openpyxl)
#
# 将 LevelingWorkbook / TraversingWorkbook 输出为 .xlsx 工作簿.
# 样式: 表头行加粗 + 浅灰背景, 列宽按内容自适应.

from __future__ import annotations

import math
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models.leveling import LevelingWorkbook
from ..models.traversing import TraversingWorkbook
from ..models.common import LevelingGrade, RodType
from ._utils import (
    rad_to_dms, format_meter, format_mm, format_arcsec, build_disclaimer,
    build_per_face_direction_values, _TWO_PI, _ARCSEC_PER_RAD,
)


# ──────────────────────────────────────────────────────────────────────
# 样式
# ──────────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9",
                           fill_type="solid")


def _write_header_row(ws, row_num: int, headers: List[str]):
    """写入表头行 (加粗 + 灰底)."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _write_data_row(ws, row_num: int, values: List):
    """写入数据行."""
    for col, v in enumerate(values, 1):
        ws.cell(row=row_num, column=col, value=v)


def _auto_column_width(ws):
    """根据内容自适应列宽 (简易版)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                # 粗略估计: 中文字符算 2 宽度
                width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, width)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


# ──────────────────────────────────────────────────────────────────────
# 精度
# ──────────────────────────────────────────────────────────────────────

_READING_DP = {
    LevelingGrade.GRADE_2: 4, LevelingGrade.GRADE_3: 3,
    LevelingGrade.GRADE_4: 3, LevelingGrade.EXTRA: 3,
}
_HEIGHT_DIFF_DP = {
    LevelingGrade.GRADE_2: 5, LevelingGrade.GRADE_3: 4,
    LevelingGrade.GRADE_4: 4, LevelingGrade.EXTRA: 3,
}


# ──────────────────────────────────────────────────────────────────────
# 水准手簿 Excel
# ──────────────────────────────────────────────────────────────────────

def _leveling_metadata_sheet(wb_excel: Workbook, workbook: LevelingWorkbook):
    """Sheet 1: 表头元数据."""
    ws = wb_excel.active
    ws.title = "表头"

    rows = [
        ("项目", workbook.project_name or "-"),
        ("等级", workbook.grade.value),
    ]

    if workbook.sections:
        md = workbook.sections[0].metadata
        rows.extend([
            ("日期", md.date),
            ("天气", md.weather or "-"),
            ("仪器", f"{md.instrument_model} ({md.instrument_serial})"),
            ("观测者", md.observer),
            ("记录者", md.recorder),
        ])
        route = workbook.sections[0].route
        rows.extend([
            ("起点", f"{route.start_point_name} (H={route.start_point_height:.3f}m)"),
            ("终点", f"{route.end_point_name} (H={route.end_point_height:.3f}m)"),
        ])
    elif workbook.extra_sections:
        md = workbook.extra_sections[0].metadata
        rows.extend([
            ("日期", md.date),
            ("天气", md.weather or "-"),
            ("仪器", f"{md.instrument_model} ({md.instrument_serial})"),
            ("观测者", md.observer),
            ("记录者", md.recorder),
        ])
        route = workbook.extra_sections[0].route
        rows.extend([
            ("起点", f"{route.start_point_name} (H={route.start_point_height:.3f}m)"),
            ("终点", f"{route.end_point_name} (H={route.end_point_height:.3f}m)"),
        ])

    rows.append(("", ""))
    rows.append(("教学声明", build_disclaimer(workbook)))

    _write_header_row(ws, 1, ["项目", "内容"])
    for i, (k, v) in enumerate(rows, 2):
        _write_data_row(ws, i, [k, v])

    _auto_column_width(ws)


def _leveling_observation_sheet(wb_excel: Workbook, workbook: LevelingWorkbook):
    """Sheet 2: 观测记录."""
    ws = wb_excel.create_sheet("观测记录")

    for sec in workbook.sections:
        grade = sec.grade
        rdp = _READING_DP.get(grade, 3)
        hdp = _HEIGHT_DIFF_DP.get(grade, 4)
        is_df = (sec.rod_back.rod_type == RodType.DOUBLE_FACE)

        if is_df:
            headers = [
                "站号", "后视点", "前视点",
                "后上丝", "后下丝", "后黑中", "后红中",
                "前上丝", "前下丝", "前黑中", "前红中",
                "后视距", "前视距", "视距差", "累积差",
                "K+黑-红(后)", "K+黑-红(前)",
                "h黑", "h红", "h中",
            ]
        else:
            headers = [
                "站号", "后视点", "前视点",
                "后上丝", "后下丝", "后基中", "后辅中",
                "前上丝", "前下丝", "前基中", "前辅中",
                "后视距", "前视距", "视距差", "累积差",
                "基辅差(后)", "基辅差(前)",
                "h基", "h辅", "h中",
            ]

        _write_header_row(ws, 1, headers)
        for i, st in enumerate(sec.stations, 2):
            bs = st.backsight
            fs = st.foresight
            vals = [
                st.station_number,
                st.backsight_point, st.foresight_point,
                bs.upper_wire_m, bs.lower_wire_m,
                bs.black_mid_m,
                bs.red_mid_m if is_df else bs.aux_mid_m,
                fs.upper_wire_m, fs.lower_wire_m,
                fs.black_mid_m,
                fs.red_mid_m if is_df else fs.aux_mid_m,
                st.stadia_back_m, st.stadia_fore_m,
                st.distance_diff_m, st.cumulative_diff_m,
                st.k_plus_black_minus_red_back_mm if is_df
                else st.base_aux_reading_diff_back_mm,
                st.k_plus_black_minus_red_fore_mm if is_df
                else st.base_aux_reading_diff_fore_mm,
                st.height_diff_black_m if is_df else st.height_diff_basic_m,
                st.height_diff_red_m if is_df else st.height_diff_aux_m,
                st.height_diff_mean_m,
            ]
            _write_data_row(ws, i, vals)

    # 等外
    for sec in workbook.extra_sections:
        headers = [
            "站号", "后视点", "前视点",
            "后视1", "前视1", "h1",
            "后视2", "前视2", "h2",
            "|h1-h2|(mm)", "h中",
        ]
        _write_header_row(ws, 1, headers)
        for i, st in enumerate(sec.stations, 2):
            _write_data_row(ws, i, [
                st.station_number,
                st.backsight_point, st.foresight_point,
                st.backsight_1_m, st.foresight_1_m, st.height_diff_1_m,
                st.backsight_2_m, st.foresight_2_m, st.height_diff_2_m,
                st.height_diff_diff_mm, st.height_diff_mean_m,
            ])

    _auto_column_width(ws)


def _leveling_summary_sheet(wb_excel: Workbook, workbook: LevelingWorkbook):
    """Sheet 3: 路线汇总 + 合规."""
    ws = wb_excel.create_sheet("汇总")

    row = 1
    _write_header_row(ws, row, ["检核项", "计算值", "限差", "通过"])
    row += 1

    for sec in workbook.sections + workbook.extra_sections:
        items = []
        if sec.sum_backsight_m is not None:
            items.append(("SUM(后视)", f"{sec.sum_backsight_m:.6f} m", "", ""))
        if sec.sum_foresight_m is not None:
            items.append(("SUM(前视)", f"{sec.sum_foresight_m:.6f} m", "", ""))
        if sec.sum_height_diff_m is not None:
            items.append(("SUM(高差)", f"{sec.sum_height_diff_m:.6f} m", "", ""))
        if sec.total_distance_km is not None:
            items.append(("路线总长", f"{sec.total_distance_km:.3f} km", "", ""))
        if sec.closure_error_mm is not None:
            limit_str = f"±{sec.closure_limit_mm:.1f} mm" if sec.closure_limit_mm else ""
            passed = "✓" if sec.closure_limit_mm and abs(sec.closure_error_mm) <= sec.closure_limit_mm + 0.01 else ""
            items.append(("闭合差", f"{sec.closure_error_mm:.3f} mm", limit_str, passed))

        for item in items:
            _write_data_row(ws, row, list(item))
            row += 1

    _auto_column_width(ws)


# ──────────────────────────────────────────────────────────────────────
# 导线手簿 Excel
# ──────────────────────────────────────────────────────────────────────

def _traversing_metadata_sheet(wb_excel: Workbook, workbook: TraversingWorkbook):
    """Sheet 1: 表头."""
    ws = wb_excel.active
    ws.title = "表头"

    rows = [
        ("项目", workbook.project_name or "-"),
        ("等级", workbook.grade.value),
        ("仪器等级", workbook.instrument_grade.value),
    ]
    if workbook.metadata:
        md = workbook.metadata
        rows.extend([
            ("日期", md.date),
            ("天气", md.weather or "-"),
            ("仪器", f"{md.instrument_model} ({md.instrument_serial})"),
            ("观测者", md.observer),
            ("记录者", md.recorder),
        ])
    if workbook.info:
        info = workbook.info
        rows.extend([
            ("导线名", info.name),
            ("起点", f"{info.start_point_name} ({info.start_point_x:.3f}, {info.start_point_y:.3f})"),
            ("终点", f"{info.end_point_name} ({info.end_point_x:.3f}, {info.end_point_y:.3f})"),
            ("角度定义", info.angle_definition.value),
        ])
        if info.start_reference_azimuth is not None:
            ref_s = f"{info.start_reference_point or '?'}→{info.start_point_name}"
            ref_e = f"{info.end_point_name}→{info.end_reference_point or '?'}"
            rows.extend([
                ("方位基准", f"{ref_s} (起始), {ref_e} (终止)"),
                ("起始方位角", f"{ref_s} = {rad_to_dms(info.start_reference_azimuth)}"),
                ("终止方位角", f"{ref_e} = {rad_to_dms(info.end_reference_azimuth)}"),
            ])

    rows.append(("", ""))
    rows.append(("教学声明", build_disclaimer(workbook)))

    _write_header_row(ws, 1, ["项目", "内容"])
    for i, (k, v) in enumerate(rows, 2):
        _write_data_row(ws, i, [k, v])
    _auto_column_width(ws)


def _traversing_angle_sheet(wb_excel: Workbook, workbook: TraversingWorkbook):
    """Sheet 2: 角度观测.

    格式规则:
    - 后视行: 仅填读数和 2C, 方向值和归零方向值留空
    - 前视L行: 填读数、2C、方向值(前视L-后视L), 归零方向值留空
    - 前视R行: 填读数、2C、方向值(前视R-后视R)、归零方向值(均值)
    """
    ws = wb_excel.create_sheet("角度观测")

    headers = ["测站", "测回", "目标", "盘位", "读数(DMS)", "2C(\")",
               "方向值", "归零方向值"]
    _write_header_row(ws, 1, headers)

    row = 2
    for obs in workbook.angle_observations:
        for aset in obs.sets:
            per_face_dv = build_per_face_direction_values(
                aset, obs.backsight_target)

            # 按 L/R 分组收集前视盘位方向值
            foresight_dv_groups = {}
            for (tgt, face), dv in per_face_dv.items():
                foresight_dv_groups.setdefault(tgt, {})[face] = dv

            for dr in aset.directions:
                two_c = aset.two_c_values_rad.get(dr.target)
                is_backsight = (dr.target == obs.backsight_target)

                dv_val = per_face_dv.get((dr.target, dr.face.value))
                dv_str = rad_to_dms(dv_val) if dv_val is not None else ""

                # 归零方向值: 仅前视R行填写
                zdr_str = ""
                if (not is_backsight
                        and dr.face.value == "R"
                        and dr.target in foresight_dv_groups):
                    dv_group = foresight_dv_groups[dr.target]
                    if "L" in dv_group and "R" in dv_group:
                        zdr = (dv_group["L"] + dv_group["R"]) / 2.0
                        zdr = zdr % _TWO_PI
                        zdr_str = rad_to_dms(zdr)

                _write_data_row(ws, row, [
                    obs.station_name,
                    aset.set_number,
                    dr.target,
                    dr.face.value,
                    rad_to_dms(dr.reading_rad),
                    f"{two_c * _ARCSEC_PER_RAD:.2f}" if two_c is not None else "",
                    "" if is_backsight else dv_str,
                    zdr_str,
                ])
                row += 1

    _auto_column_width(ws)


def _traversing_distance_sheet(wb_excel: Workbook, workbook: TraversingWorkbook):
    """Sheet 3: 距离观测."""
    ws = wb_excel.create_sheet("距离观测")

    headers = ["边名", "方向", "读数1(m)", "读数2(m)", "读数3(m)",
               "读数差(mm)", "均值(m)", "最终距离(m)"]
    _write_header_row(ws, 1, headers)

    row = 2
    for edge in workbook.distance_observations:
        for label, sets in [("往测", edge.forward_sets),
                            ("返测", edge.backward_sets)]:
            for ds in sets:
                vals = [r.reading_m for r in ds.readings]
                _write_data_row(ws, row, [
                    edge.edge_name,
                    label,
                    vals[0] if len(vals) > 0 else None,
                    vals[1] if len(vals) > 1 else None,
                    vals[2] if len(vals) > 2 else None,
                    ds.reading_diff_mm,
                    ds.mean_distance_m,
                    edge.final_distance_m,
                ])
                row += 1

    _auto_column_width(ws)


def _traversing_computation_sheet(wb_excel: Workbook, workbook: TraversingWorkbook):
    """Sheet 4: 成果计算."""
    if workbook.computation is None:
        return

    ws = wb_excel.create_sheet("成果计算")
    comp = workbook.computation

    headers = ["点名", "观测角(DMS)", "方位角(DMS)", "距离(m)",
               "Δx(m)", "Δy(m)", "X(m)", "Y(m)"]
    _write_header_row(ws, 1, headers)

    row = 2
    n_edges = len(comp.edge_records)
    n_points = len(comp.point_records)

    for i in range(max(n_points, n_edges)):
        if i < n_points:
            pr = comp.point_records[i]
            _write_data_row(ws, row, [
                pr.point_name,
                rad_to_dms(pr.observed_angle_rad) if pr.observed_angle_rad else "",
                "", "", "", "",
                pr.x_m, pr.y_m,
            ])
            row += 1
        if i < n_edges:
            er = comp.edge_records[i]
            _write_data_row(ws, row, [
                f"→{er.point_name}",
                rad_to_dms(er.observed_angle_rad) if er.observed_angle_rad else "",
                rad_to_dms(er.azimuth_rad) if er.azimuth_rad else "",
                er.distance_m,
                er.delta_x_m, er.delta_y_m,
                "", "",
            ])
            row += 1

    # 闭合差
    row += 1
    _write_header_row(ws, row, ["闭合差项", "值"])
    row += 1
    closure_items = [
        ("方位角闭合差(\")", comp.azimuth_closure_error_arcsec),
        ("f_x(m)", comp.fx_m),
        ("f_y(m)", comp.fy_m),
        ("f_D(m)", comp.fd_m),
        ("全长(m)", comp.total_length_m),
    ]
    for name, val in closure_items:
        _write_data_row(ws, row, [name, val])
        row += 1

    _auto_column_width(ws)


# ──────────────────────────────────────────────────────────────────────
# 主入口
# ──────────────────────────────────────────────────────────────────────

def workbook_to_excel(workbook, filepath: str):
    """
    将手簿输出为 Excel 工作簿 (.xlsx).

    参数:
        workbook: LevelingWorkbook 或 TraversingWorkbook
        filepath: 输出文件路径 (.xlsx)
    """
    wb = Workbook()

    if isinstance(workbook, LevelingWorkbook):
        _leveling_metadata_sheet(wb, workbook)
        _leveling_observation_sheet(wb, workbook)
        _leveling_summary_sheet(wb, workbook)
    elif isinstance(workbook, TraversingWorkbook):
        _traversing_metadata_sheet(wb, workbook)
        _traversing_angle_sheet(wb, workbook)
        _traversing_distance_sheet(wb, workbook)
        _traversing_computation_sheet(wb, workbook)

    wb.save(filepath)
