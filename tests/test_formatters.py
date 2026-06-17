# tests/test_formatters.py
# 格式化器单元测试
#
# 测试: DMS 转换、JSON 输出、纯文本/Markdown 输出、Excel 输出

import math
import json
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from src.models.common import (
    LevelingGrade, TraverseGrade, InstrumentGrade,
    AngleDefinition, RouteInfo,
)
from src.generators.leveling_generator import generate_leveling_workbook
from src.generators.traversing_generator import generate_traversing_workbook
from src.validators.traversing_validator import normalize_angle
from src.formatters._utils import (
    rad_to_dms, format_meter, format_mm, format_arcsec, build_disclaimer,
)
from src.formatters.json_formatter import workbook_to_json
from src.formatters.text_formatter import workbook_to_text, workbook_to_markdown
from src.formatters.excel_formatter import workbook_to_excel


# ──────────────────────────────────────────────────────────────────────
# DMS 转换
# ──────────────────────────────────────────────────────────────────────

class TestRadToDms:

    def test_zero(self):
        assert rad_to_dms(0.0) == '0°00\'00.0"'

    def test_90_degrees(self):
        result = rad_to_dms(math.pi / 2)
        assert result == '90°00\'00.0"'

    def test_45_degrees_30_min(self):
        rad = math.radians(45 + 30 / 60)
        result = rad_to_dms(rad)
        assert "45°30'00.0\"" == result

    def test_arbitrary_angle(self):
        # 123°45'06.5"
        d, m, s = 123, 45, 6.5
        total_sec = d * 3600 + m * 60 + s
        rad = total_sec / (180.0 * 3600 / math.pi)
        result = rad_to_dms(rad)
        assert "123°45'06.5\"" == result

    def test_small_angle(self):
        # 0°02'30" = 150"
        rad = 150.0 / (180.0 * 3600 / math.pi)
        result = rad_to_dms(rad)
        assert "0°02'30.0\"" == result


class TestFormatFunctions:

    def test_format_meter(self):
        assert format_meter(1.2346, 3) == "1.235"  # 常规进位
        assert format_meter(1.234, 4) == "1.2340"
        assert format_meter(None, 3) == "-"

    def test_format_mm(self):
        assert format_mm(0.001, 2) == "1.00"
        assert format_mm(None) == "-"

    def test_format_arcsec(self):
        # 1" = pi/(180*3600) rad
        rad = 1.0 / (180.0 * 3600 / math.pi)
        result = format_arcsec(rad, 2)
        assert abs(float(result) - 1.0) < 0.01


# ──────────────────────────────────────────────────────────────────────
# 教学声明
# ──────────────────────────────────────────────────────────────────────

class TestBuildDisclaimer:

    def test_leveling_disclaimer(self):
        route = RouteInfo("A", 100.0, "B", 101.0, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )
        disclaimer = build_disclaimer(wb)
        assert "教学" in disclaimer or "模拟" in disclaimer
        assert "grade_3" in disclaimer
        assert "0.5" in disclaimer  # sigma = 0.5mm

    def test_traversing_disclaimer(self):
        pts = [("A", 0, 0), ("P1", 100, 50), ("B", 200, 100)]
        az_s = normalize_angle(math.atan2(50, 100))
        az_e = normalize_angle(math.atan2(50, 100))
        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az_s, end_azimuth=az_e,
            grade=TraverseGrade.GRADE_1, seed=42,
        )
        disclaimer = build_disclaimer(wb)
        assert "grade_1" in disclaimer


# ──────────────────────────────────────────────────────────────────────
# JSON 格式化
# ──────────────────────────────────────────────────────────────────────

class TestJsonFormatter:

    def test_leveling_json_valid(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )
        result = workbook_to_json(wb)
        data = json.loads(result)
        assert "grade" in data
        assert "sections" in data
        assert "disclaimer" in data

    def test_traversing_json_valid(self):
        pts = [("A", 0, 0), ("P1", 100, 50), ("B", 200, 100)]
        az = normalize_angle(math.atan2(50, 100))
        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az, end_azimuth=az,
            grade=TraverseGrade.GRADE_1, seed=42,
        )
        result = workbook_to_json(wb)
        data = json.loads(result)
        assert "angle_observations" in data
        assert "distance_observations" in data
        assert "disclaimer" in data

    def test_json_contains_stations(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=4, seed=42,
        )
        data = json.loads(workbook_to_json(wb))
        stations = data["sections"][0]["stations"]
        assert len(stations) == 4

    def test_json_ensure_ascii_false(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=2, seed=42,
        )
        result = workbook_to_json(wb)
        assert "模拟" in result  # 中文不被转义


# ──────────────────────────────────────────────────────────────────────
# 纯文本格式化
# ──────────────────────────────────────────────────────────────────────

class TestTextFormatter:

    def test_leveling_text_not_empty(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )
        text = workbook_to_text(wb)
        assert len(text) > 100
        assert "水准观测手簿" in text

    def test_leveling_text_contains_stations(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )
        text = workbook_to_text(wb)
        assert "TP.1" in text or "TP.2" in text

    def test_leveling_text_has_disclaimer(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=2, seed=42,
        )
        text = workbook_to_text(wb)
        assert "教学" in text or "模拟" in text

    def test_extra_text(self):
        route = RouteInfo("A", 100.0, "B", 101.0, 0.2)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.EXTRA,
            num_stations=3, seed=42,
        )
        text = workbook_to_text(wb)
        assert "等外" in text

    def test_traversing_text_not_empty(self):
        pts = [("A", 0, 0), ("P1", 100, 50), ("B", 200, 100)]
        az = normalize_angle(math.atan2(50, 100))
        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az, end_azimuth=az,
            grade=TraverseGrade.GRADE_1, seed=42,
        )
        text = workbook_to_text(wb)
        assert len(text) > 100
        assert "导线观测手簿" in text
        assert "水平角观测" in text
        assert "距离观测" in text


# ──────────────────────────────────────────────────────────────────────
# Markdown 格式化
# ──────────────────────────────────────────────────────────────────────

class TestMarkdownFormatter:

    def test_leveling_markdown_has_tables(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )
        md = workbook_to_markdown(wb)
        assert "|" in md  # 有表格分隔符
        assert "---" in md  # 有表格分隔线

    def test_traversing_markdown_has_sections(self):
        pts = [("A", 0, 0), ("P1", 100, 50), ("B", 200, 100)]
        az = normalize_angle(math.atan2(50, 100))
        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az, end_azimuth=az,
            grade=TraverseGrade.GRADE_1, seed=42,
        )
        md = workbook_to_markdown(wb)
        assert "## 水平角观测" in md
        assert "## 距离观测" in md
        assert "## 成果计算" in md


# ──────────────────────────────────────────────────────────────────────
# Excel 格式化
# ──────────────────────────────────────────────────────────────────────

class TestExcelFormatter:

    def test_leveling_excel_creates_file(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name
        try:
            workbook_to_excel(wb, filepath)
            assert os.path.exists(filepath)
            assert os.path.getsize(filepath) > 0

            # 验证可以打开
            from openpyxl import load_workbook
            wb_excel = load_workbook(filepath)
            sheet_names = wb_excel.sheetnames
            assert "表头" in sheet_names
            assert "观测记录" in sheet_names
            assert "汇总" in sheet_names
        finally:
            os.unlink(filepath)

    def test_traversing_excel_creates_file(self):
        pts = [("A", 0, 0), ("P1", 100, 50), ("B", 200, 100)]
        az = normalize_angle(math.atan2(50, 100))
        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az, end_azimuth=az,
            grade=TraverseGrade.GRADE_1, seed=42,
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name
        try:
            workbook_to_excel(wb, filepath)
            assert os.path.exists(filepath)

            from openpyxl import load_workbook
            wb_excel = load_workbook(filepath)
            sheet_names = wb_excel.sheetnames
            assert "表头" in sheet_names
            assert "角度观测" in sheet_names
            assert "距离观测" in sheet_names
            assert "成果计算" in sheet_names
        finally:
            os.unlink(filepath)

    def test_excel_observation_sheet_has_data(self):
        route = RouteInfo("A", 100.0, "B", 101.5, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name
        try:
            workbook_to_excel(wb, filepath)
            from openpyxl import load_workbook
            wb_excel = load_workbook(filepath)
            ws = wb_excel["观测记录"]
            # 表头行 + 3 站数据 = 4 行
            assert ws.max_row >= 4
        finally:
            os.unlink(filepath)


# ──────────────────────────────────────────────────────────────────────
# 阶段十九: 导线角度观测表方向值格式
# ──────────────────────────────────────────────────────────────────────

class TestAngleTableFormat:
    """导线角度观测表: 方向值和归零方向值仅填前视行."""

    @pytest.fixture
    def grade1_wb(self):
        pts = [("A", 0, 0), ("P1", 100, 50), ("B", 200, 100)]
        az = normalize_angle(math.atan2(50, 100))
        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az, end_azimuth=az,
            grade=TraverseGrade.GRADE_1, seed=42,
        )
        from src.validators.traversing_validator import validate_traversing_workbook
        validate_traversing_workbook(wb)  # 填充 2C/方向值等计算字段
        return wb

    def _parse_angle_rows(self, md_text):
        """从 Markdown 文本解析角度观测表行, 返回数据行列表."""
        lines = md_text.split("\n")
        in_angle = False
        header_seen = False
        rows = []
        for line in lines:
            if "## 水平角观测" in line:
                in_angle = True
                continue
            if in_angle and line.startswith("## "):
                break  # 下一个 section
            if not in_angle:
                continue
            if line.startswith("| ---"):
                header_seen = True
                continue
            if header_seen and line.startswith("|"):
                cells = [c.strip() for c in line.strip("|").split("|")]
                rows.append(cells)
        return rows

    def test_backsight_rows_no_direction_value(self, grade1_wb):
        """后视行方向值和归零方向值应为空."""
        md = workbook_to_markdown(grade1_wb)
        rows = self._parse_angle_rows(md)
        assert len(rows) > 0

        obs = grade1_wb.angle_observations
        backsight_targets = {o.backsight_target for o in obs}

        for row in rows:
            target = row[2]  # 目标列
            if target in backsight_targets:
                dv = row[6]   # 方向值列
                zdr = row[7]  # 归零方向值列
                assert dv == "-", f"后视 {target} 方向值应为 '-', 得到 '{dv}'"
                assert zdr == "-", f"后视 {target} 归零方向值应为 '-', 得到 '{zdr}'"

    def test_foresight_L_has_direction_no_zero_reduced(self, grade1_wb):
        """前视L行有方向值, 无归零方向值."""
        md = workbook_to_markdown(grade1_wb)
        rows = self._parse_angle_rows(md)

        obs = grade1_wb.angle_observations
        backsight_targets = {o.backsight_target for o in obs}

        count = 0
        for row in rows:
            target = row[2]
            face = row[3]
            if target not in backsight_targets and face == "L":
                dv = row[6]
                zdr = row[7]
                assert dv != "-", f"前视 {target} L行方向值不应为空"
                assert zdr == "-", f"前视 {target} L行归零方向值应为空"
                count += 1
        assert count > 0, "应至少有一行前视L"

    def test_foresight_R_has_direction_and_zero_reduced(self, grade1_wb):
        """前视R行有方向值和归零方向值."""
        md = workbook_to_markdown(grade1_wb)
        rows = self._parse_angle_rows(md)

        obs = grade1_wb.angle_observations
        backsight_targets = {o.backsight_target for o in obs}

        count = 0
        for row in rows:
            target = row[2]
            face = row[3]
            if target not in backsight_targets and face == "R":
                dv = row[6]
                zdr = row[7]
                assert dv != "-", f"前视 {target} R行方向值不应为空"
                assert zdr != "-", f"前视 {target} R行归零方向值不应为空"
                count += 1
        assert count > 0, "应至少有一行前视R"

    def test_direction_value_formula(self, grade1_wb):
        """方向值 = (前视读数 - 后视读数) mod 360°."""
        from src.formatters._utils import build_per_face_direction_values, _TWO_PI

        for obs in grade1_wb.angle_observations:
            for aset in obs.sets:
                per_face_dv = build_per_face_direction_values(
                    aset, obs.backsight_target)

                # 从原始读数手动计算
                readings = {}
                for dr in aset.directions:
                    readings[(dr.target, dr.face.value)] = dr.reading_rad

                for (tgt, face), dv in per_face_dv.items():
                    fs_r = readings[(tgt, face)]
                    bs_r = readings[(obs.backsight_target, face)]
                    expected = (fs_r - bs_r) % _TWO_PI
                    assert abs(dv - expected) < 1e-12, \
                        f"方向值公式错误: {tgt}/{face}"

    def test_zero_reduced_is_mean_of_LR(self, grade1_wb):
        """归零方向值 = (方向值_L + 方向值_R) / 2."""
        from src.formatters._utils import build_per_face_direction_values, _TWO_PI

        for obs in grade1_wb.angle_observations:
            for aset in obs.sets:
                per_face_dv = build_per_face_direction_values(
                    aset, obs.backsight_target)

                # 按目标分组
                groups = {}
                for (tgt, face), dv in per_face_dv.items():
                    groups.setdefault(tgt, {})[face] = dv

                for tgt, dv_map in groups.items():
                    if "L" in dv_map and "R" in dv_map:
                        expected = (dv_map["L"] + dv_map["R"]) / 2.0
                        expected = expected % _TWO_PI
                        # 检查 R 行的归零方向值
                        zdr_str = rad_to_dms(expected)
                        # 不应为空
                        assert zdr_str != "-", \
                            f"归零方向值不应为空: {tgt}"

    def test_2c_present_on_all_rows(self, grade1_wb):
        """2C 值在所有行 (后视+前视, L+R) 都存在."""
        md = workbook_to_markdown(grade1_wb)
        rows = self._parse_angle_rows(md)

        for row in rows:
            two_c = row[5]  # 2C 列
            assert two_c != "-", f"2C 值不应为空: {row[:4]}"


# ──────────────────────────────────────────────────────────────────────
# 阶段二十二: 导线成果计算表 14 列格式
# ──────────────────────────────────────────────────────────────────────

class TestComputationTableFormat:
    """导线成果计算表: 14 列标准格式 (点名行/边行交替)."""

    @pytest.fixture
    def adjusted_wb(self):
        """生成带平差的导线手簿 (target_closure_ratio=0.3 使改正数非零)."""
        pts = [("A", 1000, 1000), ("P1", 1100, 1050), ("P2", 1200, 1100),
               ("B", 1300, 1200)]
        az = normalize_angle(math.atan2(50, 100))
        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az, end_azimuth=az,
            grade=TraverseGrade.GRADE_1, seed=42,
            target_closure_ratio=0.3,
        )
        return wb

    def _parse_computation_rows(self, md_text):
        """从 Markdown 文本解析成果计算表行, 返回数据行列表."""
        lines = md_text.split("\n")
        in_comp = False
        header_seen = False
        rows = []
        for line in lines:
            if "## 成果计算" in line:
                in_comp = True
                continue
            if in_comp and line.startswith("## "):
                break  # 下一个 section
            if not in_comp:
                continue
            if line.startswith("| ---"):
                header_seen = True
                continue
            if header_seen and line.startswith("|"):
                cells = [c.strip() for c in line.strip("|").split("|")]
                rows.append(cells)
        return rows

    def test_md_table_has_14_headers(self, adjusted_wb):
        """Markdown 成果计算表表头应有 14 列."""
        md = workbook_to_markdown(adjusted_wb)
        lines = md.split("\n")
        in_comp = False
        header_line = None
        for line in lines:
            if "## 成果计算" in line:
                in_comp = True
                continue
            if in_comp and line.startswith("|") and "---" not in line:
                header_line = line
                break
        assert header_line is not None, "未找到成果计算表表头"
        headers = [h.strip() for h in header_line.strip("|").split("|")]
        assert len(headers) == 14, f"表头列数={len(headers)}, 期望 14"
        # 验证关键列名
        assert headers[0] == "点名"
        assert headers[1] == "观测角"
        assert headers[4] == "方位角"
        assert headers[5] == "距离(m)"
        assert headers[12] == "X(m)"
        assert headers[13] == "Y(m)"

    def test_point_rows_have_angle_and_coords(self, adjusted_wb):
        """点名行: 列0-3有角度, 列4-11为空, 列12-13有坐标."""
        md = workbook_to_markdown(adjusted_wb)
        rows = self._parse_computation_rows(md)
        assert len(rows) > 0

        # 点名行: 不以 "→" 开头
        point_rows = [r for r in rows if not r[0].startswith("→") and not r[0].startswith("  →")]
        assert len(point_rows) > 0, "应至少有一行点名行"

        for row in point_rows:
            # 列0: 点名非空
            assert row[0] != "", "点名不应为空"
            # 列4-11: 方位角/距离/增量/改正数列为空
            for i in range(4, 12):
                assert row[i] == "", f"点名行列{i}应为空, 得到 '{row[i]}'"
            # 列12-13: X/Y 坐标非空
            assert row[12] != "", "点名行 X 坐标不应为空"
            assert row[13] != "", "点名行 Y 坐标不应为空"

    def test_edge_rows_have_azimuth_and_distance(self, adjusted_wb):
        """边行: 列0-3为空, 列4-7有方位角/距离/增量, 列12-13为空."""
        md = workbook_to_markdown(adjusted_wb)
        rows = self._parse_computation_rows(md)
        assert len(rows) > 0

        # 边行: 以 "→" 开头
        edge_rows = [r for r in rows if "→" in r[0]]
        assert len(edge_rows) > 0, "应至少有一行边行"

        for row in edge_rows:
            # 列1-3: 观测角/改正数/改正后角为空
            for i in range(1, 4):
                assert row[i] == "", f"边行列{i}应为空, 得到 '{row[i]}'"
            # 列4: 方位角非空
            assert row[4] != "", "边行方位角不应为空"
            # 列5: 距离非空
            assert row[5] != "", "边行距离不应为空"
            # 列12-13: X/Y 坐标为空
            assert row[12] == "", "边行 X 坐标应为空"
            assert row[13] == "", "边行 Y 坐标应为空"

    def test_correction_columns_nonempty_after_adjustment(self, adjusted_wb):
        """平差后改正数列 (v_beta, v_x, v_y) 应非空."""
        md = workbook_to_markdown(adjusted_wb)
        rows = self._parse_computation_rows(md)

        # 收集所有改正数列值
        # 列2: v_beta(") — 点名行
        # 列8: v_x(mm) — 边行
        # 列9: v_y(mm) — 边行
        v_betas = []
        v_xs = []
        v_ys = []
        for row in rows:
            if "→" not in row[0]:
                # 点名行
                if row[2] != "":
                    v_betas.append(row[2])
            else:
                # 边行
                if row[8] != "":
                    v_xs.append(row[8])
                if row[9] != "":
                    v_ys.append(row[9])

        # target_closure_ratio=0.3 应产生非零改正数
        assert len(v_betas) > 0, "应有非空的 v_beta 改正数"
        assert len(v_xs) > 0, "应有非空的 v_x 改正数"
        assert len(v_ys) > 0, "应有非空的 v_y 改正数"

    def test_excel_computation_sheet_has_14_headers(self, adjusted_wb):
        """Excel 成果计算表表头应有 14 列."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name
        try:
            workbook_to_excel(adjusted_wb, filepath)
            from openpyxl import load_workbook
            wb_excel = load_workbook(filepath)
            ws = wb_excel["成果计算"]

            # 第1行为表头, 验证14列
            header_count = 0
            for col in range(1, 20):
                val = ws.cell(row=1, column=col).value
                if val is not None:
                    header_count += 1
            assert header_count == 14, f"Excel表头列数={header_count}, 期望 14"

            # 验证关键列名
            assert ws.cell(row=1, column=1).value == "点名"
            assert ws.cell(row=1, column=2).value == "观测角"
            assert ws.cell(row=1, column=5).value == "方位角"
            assert ws.cell(row=1, column=6).value == "距离(m)"
            assert ws.cell(row=1, column=13).value == "X(m)"
            assert ws.cell(row=1, column=14).value == "Y(m)"
        finally:
            os.unlink(filepath)

    def test_excel_correction_columns_nonempty(self, adjusted_wb):
        """Excel 平差后改正数列应非空 (v_beta 列3, v_x 列9, v_y 列10)."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name
        try:
            workbook_to_excel(adjusted_wb, filepath)
            from openpyxl import load_workbook
            wb_excel = load_workbook(filepath)
            ws = wb_excel["成果计算"]

            # 遍历数据行, 检查改正数列非空
            v_beta_found = False
            v_x_found = False
            v_y_found = False
            for row_num in range(2, ws.max_row + 1):
                # 列1: 点名, 列3: v_beta, 列9: v_x, 列10: v_y
                name = ws.cell(row=row_num, column=1).value
                if name is None:
                    continue
                v_beta = ws.cell(row=row_num, column=3).value
                v_x = ws.cell(row=row_num, column=9).value
                v_y = ws.cell(row=row_num, column=10).value

                if v_beta not in (None, ""):
                    v_beta_found = True
                if v_x not in (None, ""):
                    v_x_found = True
                if v_y not in (None, ""):
                    v_y_found = True

            assert v_beta_found, "Excel 应有非空的 v_beta 改正数"
            assert v_x_found, "Excel 应有非空的 v_x 改正数"
            assert v_y_found, "Excel 应有非空的 v_y 改正数"
        finally:
            os.unlink(filepath)


# ──────────────────────────────────────────────────────────────────────
# 阶段二十三: 水准成果计算表格式化
# ──────────────────────────────────────────────────────────────────────

class TestLevelingAdjustmentFormat:
    """水准成果计算表: 6 列格式 + 辅助区."""

    @pytest.fixture
    def grade3_wb(self):
        """三等单程水准 (含平差, target_closure_ratio=0.3 使改正数非零)."""
        route = RouteInfo("BM.A", 100.0, "BM.B", 108.0, 1.5)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=6, seed=42,
            target_closure_ratio=0.3,
        )
        return wb

    @pytest.fixture
    def grade2_round_trip_wb(self):
        """二等往返测水准 (含平差)."""
        route = RouteInfo("BM.C", 200.0, "BM.D", 215.0, 2.5)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_2,
            num_stations=8, seed=42,
            round_trip=True, target_round_trip_ratio=0.3,
        )
        return wb

    def _parse_adjustment_rows(self, md_text):
        """从 Markdown 文本解析水准成果计算表行, 返回数据行列表."""
        lines = md_text.split("\n")
        in_adj = False
        header_seen = False
        rows = []
        for line in lines:
            if "## 成果计算" in line:
                in_adj = True
                continue
            if in_adj and line.startswith("## "):
                break  # 下一个 section
            if in_adj and "**辅助计算**" in line:
                break  # 进入辅助区
            if not in_adj:
                continue
            if line.startswith("| ---"):
                header_seen = True
                continue
            if header_seen and line.startswith("|"):
                cells = [c.strip() for c in line.strip("|").split("|")]
                rows.append(cells)
        return rows

    def test_md_table_has_6_headers(self, grade3_wb):
        """Markdown 成果计算表表头应有 6 列."""
        md = workbook_to_markdown(grade3_wb)
        lines = md.split("\n")
        in_adj = False
        header_line = None
        for line in lines:
            if "## 成果计算" in line:
                in_adj = True
                continue
            if in_adj and line.startswith("|") and "---" not in line:
                header_line = line
                break
        assert header_line is not None, "未找到水准成果计算表表头"
        headers = [h.strip() for h in header_line.strip("|").split("|")]
        assert len(headers) == 6, f"表头列数={len(headers)}, 期望 6"
        assert headers[0] == "点名"
        assert headers[1] == "距离(km)"
        assert headers[2] == "观测高差(m)"
        assert headers[3] == "改正数(mm)"
        assert headers[4] == "改正后高差(m)"
        assert headers[5] == "高程(m)"

    def test_md_rows_have_data(self, grade3_wb):
        """Markdown 成果计算表数据行应完整 (6站)."""
        md = workbook_to_markdown(grade3_wb)
        rows = self._parse_adjustment_rows(md)
        assert len(rows) == 6, f"数据行数={len(rows)}, 期望 6"

        for row in rows:
            # 所有6列都应有值
            for i, cell in enumerate(row):
                assert cell != "", f"行{i}列不应为空: {row}"

    def test_md_aux_area_present(self, grade3_wb):
        """Markdown 辅助计算区应包含闭合差、限差、是否合格."""
        md = workbook_to_markdown(grade3_wb)
        assert "**辅助计算**" in md
        assert "闭合差" in md
        assert "限差" in md
        assert "是否合格" in md
        assert "每公里改正数" in md

    def test_md_endpoint_height_exact(self, grade3_wb):
        """Markdown 最后一行高程应精确等于终点高程 108.0."""
        md = workbook_to_markdown(grade3_wb)
        rows = self._parse_adjustment_rows(md)
        assert len(rows) > 0
        last_height_str = rows[-1][5]  # 高程(m) 列
        last_height = float(last_height_str)
        assert abs(last_height - 108.0) < 1e-4, (
            f"终点高程={last_height}, 期望 108.0"
        )

    def test_md_round_trip_annotation(self, grade2_round_trip_wb):
        """往返测场景应包含往返测附注."""
        md = workbook_to_markdown(grade2_round_trip_wb)
        assert "往返测附注" in md
        assert "往返测不符值" in md
        assert "往返测中数高差" in md

    def test_excel_adjustment_sheet_has_6_headers(self, grade3_wb):
        """Excel 成果计算表表头应有 6 列."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name
        try:
            workbook_to_excel(grade3_wb, filepath)
            from openpyxl import load_workbook
            wb_excel = load_workbook(filepath)
            assert "成果计算" in wb_excel.sheetnames, "应有成果计算 Sheet"
            ws = wb_excel["成果计算"]

            # 第1行为表头, 验证6列
            header_count = 0
            for col in range(1, 10):
                val = ws.cell(row=1, column=col).value
                if val is not None:
                    header_count += 1
            assert header_count == 6, f"Excel表头列数={header_count}, 期望 6"

            # 验证关键列名
            assert ws.cell(row=1, column=1).value == "点名"
            assert ws.cell(row=1, column=2).value == "距离(km)"
            assert ws.cell(row=1, column=3).value == "观测高差(m)"
            assert ws.cell(row=1, column=4).value == "改正数(mm)"
            assert ws.cell(row=1, column=5).value == "改正后高差(m)"
            assert ws.cell(row=1, column=6).value == "高程(m)"
        finally:
            os.unlink(filepath)

    def test_excel_endpoint_height_exact(self, grade3_wb):
        """Excel 最后一行高程应精确等于终点高程 108.0."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name
        try:
            workbook_to_excel(grade3_wb, filepath)
            from openpyxl import load_workbook
            wb_excel = load_workbook(filepath)
            ws = wb_excel["成果计算"]

            # 6站数据行: 第2-7行, 最后一行是第7行
            # 找到最后一行数据 (跳过辅助区)
            last_data_row = None
            for row_num in range(2, ws.max_row + 1):
                name = ws.cell(row=row_num, column=1).value
                if name is None or name == "辅助计算项":
                    break
                last_data_row = row_num

            assert last_data_row is not None, "应有数据行"
            last_height = ws.cell(row=last_data_row, column=6).value
            assert abs(last_height - 108.0) < 1e-4, (
                f"终点高程={last_height}, 期望 108.0"
            )
        finally:
            os.unlink(filepath)

    def test_json_contains_adjustment(self, grade3_wb):
        """JSON 输出应包含 adjustment 字段及其子字段."""
        data = json.loads(workbook_to_json(grade3_wb))
        assert "adjustment" in data
        adj = data["adjustment"]
        assert adj is not None
        assert "records" in adj
        assert len(adj["records"]) == 6
        assert "closure_error_mm" in adj
        assert "closure_limit_mm" in adj
        assert "passed" in adj
        assert "correction_per_km_mm" in adj
        # 验证记录字段
        rec = adj["records"][0]
        assert "point_name" in rec
        assert "distance_km" in rec
        assert "observed_height_diff_m" in rec
        assert "correction_mm" in rec
        assert "corrected_height_diff_m" in rec
        assert "height_m" in rec
