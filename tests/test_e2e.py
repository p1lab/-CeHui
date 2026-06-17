# tests/test_e2e.py
# 端到端集成测试
#
# 完整管线: 坐标 → 生成 → 验证 → 合规 → 四种格式输出
# 验证: 生成通过 → 验证通过 → 合规通过 → 输出非空且含教学声明

import json
import math
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from src.models.common import (
    LevelingGrade, TraverseGrade, InstrumentGrade,
    AngleDefinition, RouteInfo, SurveyMetadata,
)
from src.models.leveling import RodSpec
from src.models.common import RodType
from src.generators.leveling_generator import generate_leveling_workbook
from src.generators.traversing_generator import generate_traversing_workbook
from src.validators.leveling_validator import validate_leveling_workbook
from src.validators.traversing_validator import (
    validate_traversing_workbook, normalize_angle,
)
from src.checkers.leveling_compliance import check_leveling_compliance
from src.checkers.traversing_compliance import check_traversing_compliance
from src.formatters.json_formatter import workbook_to_json
from src.formatters.text_formatter import workbook_to_text, workbook_to_markdown
from src.formatters.excel_formatter import workbook_to_excel


# ──────────────────────────────────────────────────────────────────────
# 辅助
# ──────────────────────────────────────────────────────────────────────

def _azimuth(x1, y1, x2, y2):
    return normalize_angle(math.atan2(y2 - y1, x2 - x1))


# ──────────────────────────────────────────────────────────────────────
# 水准三等 E2E
# ──────────────────────────────────────────────────────────────────────

class TestLevelingE2E:
    """水准完整管线: 生成 → 验证 → 合规 → 四种格式输出"""

    @pytest.fixture
    def grade3_wb(self):
        route = RouteInfo("BM.A", 100.000, "BM.B", 101.500, 0.4)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=4, seed=42,
        )
        return wb

    def test_pipeline_passes(self, grade3_wb):
        """生成 → 验证通过 → 合规通过"""
        val_result = validate_leveling_workbook(grade3_wb)
        assert val_result.all_passed, "正向验证未通过"

        comp_report = check_leveling_compliance(grade3_wb)
        assert comp_report.passed, "合规检核未通过"

    def test_json_output(self, grade3_wb):
        """JSON 输出可解析, 含关键字段"""
        j = workbook_to_json(grade3_wb)
        data = json.loads(j)

        assert "sections" in data or "extra_sections" in data
        assert "disclaimer" in data
        assert "教学" in data["disclaimer"] or "模拟" in data["disclaimer"]

    def test_text_output(self, grade3_wb):
        """纯文本输出非空, 含表头和观测数据"""
        text = workbook_to_text(grade3_wb)
        assert len(text) > 100
        assert "水准" in text or "观测" in text or "站" in text
        assert "教学" in text or "模拟" in text

    def test_markdown_output(self, grade3_wb):
        """Markdown 输出含表格标记"""
        md = workbook_to_markdown(grade3_wb)
        assert "|" in md  # 表格管道符
        assert "教学" in md or "模拟" in md

    def test_excel_output(self, grade3_wb, tmp_path):
        """Excel 输出可创建, Sheet 数量正确"""
        fp = str(tmp_path / "leveling_e2e.xlsx")
        workbook_to_excel(grade3_wb, fp)

        assert os.path.exists(fp)
        assert os.path.getsize(fp) > 100

        from openpyxl import load_workbook
        xlsx = load_workbook(fp)
        # 水准应有: 表头, 观测记录, 汇总 (+ 可选合规)
        assert len(xlsx.sheetnames) >= 3


# ──────────────────────────────────────────────────────────────────────
# 水准二等 E2E (因瓦基辅)
# ──────────────────────────────────────────────────────────────────────

class TestLevelingGrade2E2E:
    """二等因瓦尺完整管线"""

    def test_grade2_pipeline(self):
        route = RouteInfo("BM.C", 50.000, "BM.D", 50.800, 0.3)
        rod_back = RodSpec("No.1", RodType.INVAR_BASIC_AUX, c_aux_m=3.0155)
        rod_fore = RodSpec("No.2", RodType.INVAR_BASIC_AUX, c_aux_m=3.0155)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_2,
            num_stations=3, rod_back=rod_back, rod_fore=rod_fore,
            seed=42,
        )

        val_result = validate_leveling_workbook(wb)
        assert val_result.all_passed

        comp_report = check_leveling_compliance(wb)
        assert comp_report.passed

        # 四种格式输出
        j = workbook_to_json(wb)
        assert json.loads(j) is not None

        text = workbook_to_text(wb)
        assert "基" in text or "辅" in text or "因瓦" in text or len(text) > 100

        md = workbook_to_markdown(wb)
        assert "|" in md

        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "grade2.xlsx")
            workbook_to_excel(wb, fp)
            assert os.path.exists(fp)


# ──────────────────────────────────────────────────────────────────────
# 水准等外 E2E
# ──────────────────────────────────────────────────────────────────────

class TestLevelingExtraE2E:
    """等外水准 (变动仪高法) 完整管线"""

    def test_extra_pipeline(self):
        route = RouteInfo("BM.E", 100.000, "BM.F", 101.000, 0.2)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.EXTRA,
            num_stations=3, seed=42,
        )

        val_result = validate_leveling_workbook(wb)
        assert val_result.all_passed

        comp_report = check_leveling_compliance(wb)
        assert comp_report.passed

        # JSON
        data = json.loads(workbook_to_json(wb))
        assert "extra_sections" in data

        # Text — 变动仪高法
        text = workbook_to_text(wb)
        assert len(text) > 50

        # Markdown
        md = workbook_to_markdown(wb)
        assert len(md) > 50


# ──────────────────────────────────────────────────────────────────────
# 导线一级 E2E
# ──────────────────────────────────────────────────────────────────────

class TestTraversingE2E:
    """导线完整管线: 坐标 → 生成 → 验证 → 合规 → 四种格式输出"""

    @pytest.fixture
    def grade1_wb(self):
        pts = [
            ("A", 1000.0, 1000.0),
            ("P1", 1100.0, 1050.0),
            ("P2", 1200.0, 1100.0),
            ("B", 1300.0, 1200.0),
        ]
        az_start = _azimuth(1000, 1000, 1100, 1050)
        az_end = _azimuth(1200, 1100, 1300, 1200)

        wb = generate_traversing_workbook(
            points=pts,
            start_azimuth=az_start,
            end_azimuth=az_end,
            grade=TraverseGrade.GRADE_1,
            num_angle_sets=2,
            seed=42,
        )
        return wb

    def test_pipeline_passes(self, grade1_wb):
        """生成 → 验证通过 → 合规通过"""
        val_result = validate_traversing_workbook(grade1_wb)
        assert val_result.all_passed, "正向验证未通过"

        comp_report = check_traversing_compliance(grade1_wb)
        assert comp_report.passed, "合规检核未通过"

    def test_json_output(self, grade1_wb):
        """JSON 输出可解析, 含角度和距离观测"""
        j = workbook_to_json(grade1_wb)
        data = json.loads(j)

        assert "angle_observations" in data
        assert "distance_observations" in data
        assert "disclaimer" in data

    def test_text_output(self, grade1_wb):
        """纯文本输出含导线信息"""
        text = workbook_to_text(grade1_wb)
        assert len(text) > 100
        assert "导线" in text or "角度" in text or "距离" in text

    def test_markdown_output(self, grade1_wb):
        """Markdown 输出含表格"""
        md = workbook_to_markdown(grade1_wb)
        assert "|" in md

    def test_excel_output(self, grade1_wb, tmp_path):
        """Excel 输出 Sheet 数量正确"""
        fp = str(tmp_path / "traverse_e2e.xlsx")
        workbook_to_excel(grade1_wb, fp)

        assert os.path.exists(fp)

        from openpyxl import load_workbook
        xlsx = load_workbook(fp)
        # 导线应有: 表头, 角度观测, 距离观测, 成果计算
        assert len(xlsx.sheetnames) >= 4


# ──────────────────────────────────────────────────────────────────────
# 导线二级 E2E
# ──────────────────────────────────────────────────────────────────────

class TestTraversingGrade2E2E:

    def test_grade2_pipeline(self):
        pts = [
            ("K1", 500.0, 500.0),
            ("T1", 600.0, 550.0),
            ("T2", 700.0, 500.0),
            ("K2", 800.0, 600.0),
        ]
        az_start = _azimuth(500, 500, 600, 550)
        az_end = _azimuth(700, 500, 800, 600)

        wb = generate_traversing_workbook(
            points=pts,
            start_azimuth=az_start,
            end_azimuth=az_end,
            grade=TraverseGrade.GRADE_2,
            num_angle_sets=1,
            seed=42,
        )

        val_result = validate_traversing_workbook(wb)
        assert val_result.all_passed

        comp_report = check_traversing_compliance(wb)
        assert comp_report.passed

        # 四种格式
        j = workbook_to_json(wb)
        assert json.loads(j) is not None

        text = workbook_to_text(wb)
        assert len(text) > 100

        md = workbook_to_markdown(wb)
        assert "|" in md


# ──────────────────────────────────────────────────────────────────────
# 导线图根 E2E
# ──────────────────────────────────────────────────────────────────────

class TestTraversingRootE2E:

    def test_root_pipeline(self):
        pts = [
            ("G1", 0.0, 0.0),
            ("R1", 50.0, 30.0),
            ("R2", 100.0, 10.0),
            ("G2", 150.0, 50.0),
        ]
        az_start = _azimuth(0, 0, 50, 30)
        az_end = _azimuth(100, 10, 150, 50)

        wb = generate_traversing_workbook(
            points=pts,
            start_azimuth=az_start,
            end_azimuth=az_end,
            grade=TraverseGrade.ROOT,
            num_angle_sets=1,
            seed=123,
        )

        val_result = validate_traversing_workbook(wb)
        assert val_result.all_passed

        comp_report = check_traversing_compliance(wb)
        assert comp_report.passed

        # 输出
        text = workbook_to_text(wb)
        assert len(text) > 100


# ──────────────────────────────────────────────────────────────────────
# 教学声明一致性
# ──────────────────────────────────────────────────────────────────────

class TestDisclaimerConsistency:
    """所有输出格式都包含教学声明"""

    def test_leveling_disclaimer_in_all(self):
        route = RouteInfo("BM.A", 100.000, "BM.B", 101.500, 0.3)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )

        j = workbook_to_json(wb)
        text = workbook_to_text(wb)
        md = workbook_to_markdown(wb)

        keyword = "教学"
        assert keyword in j, "JSON 缺少教学声明"
        assert keyword in text, "Text 缺少教学声明"
        assert keyword in md, "Markdown 缺少教学声明"

    def test_traversing_disclaimer_in_all(self):
        pts = [("A", 0, 0), ("P1", 100, 50), ("B", 200, 100)]
        az_s = _azimuth(0, 0, 100, 50)
        az_e = _azimuth(100, 50, 200, 100)

        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az_s, end_azimuth=az_e,
            grade=TraverseGrade.GRADE_1, seed=42,
        )

        j = workbook_to_json(wb)
        text = workbook_to_text(wb)
        md = workbook_to_markdown(wb)

        keyword = "教学"
        assert keyword in j, "JSON 缺少教学声明"
        assert keyword in text, "Text 缺少教学声明"
        assert keyword in md, "Markdown 缺少教学声明"


# ──────────────────────────────────────────────────────────────────────
# 可复现性 E2E
# ──────────────────────────────────────────────────────────────────────

class TestReproducibilityE2E:
    """相同种子 → 完全相同的输出"""

    def test_leveling_reproducible(self):
        route = RouteInfo("BM.A", 100.000, "BM.B", 101.500, 0.3)
        wb1 = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )
        wb2 = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=3, seed=42,
        )

        j1 = workbook_to_json(wb1)
        j2 = workbook_to_json(wb2)
        assert j1 == j2, "相同种子应产生相同 JSON"

        t1 = workbook_to_text(wb1)
        t2 = workbook_to_text(wb2)
        assert t1 == t2, "相同种子应产生相同 Text"

    def test_traversing_reproducible(self):
        pts = [("A", 0, 0), ("P1", 100, 50), ("B", 200, 100)]
        az_s = _azimuth(0, 0, 100, 50)
        az_e = _azimuth(100, 50, 200, 100)

        wb1 = generate_traversing_workbook(
            points=pts, start_azimuth=az_s, end_azimuth=az_e,
            grade=TraverseGrade.GRADE_1, seed=42,
        )
        wb2 = generate_traversing_workbook(
            points=pts, start_azimuth=az_s, end_azimuth=az_e,
            grade=TraverseGrade.GRADE_1, seed=42,
        )

        j1 = workbook_to_json(wb1)
        j2 = workbook_to_json(wb2)
        assert j1 == j2, "相同种子应产生相同 JSON"


# ──────────────────────────────────────────────────────────────────────
# 阶段二十四: 含平差的 E2E 测试
# ──────────────────────────────────────────────────────────────────────

class TestTraversingAdjustmentE2E:
    """导线 E2E (含平差): 生成 → 平差 → 验证改正后坐标闭合."""

    def _make_traverse(self, grade, seed=42, target_closure_ratio=0.0,
                       start_ref=None, end_ref=None):
        """生成导线 workbook (含平差)."""
        pts = [
            ("A", 1000.0, 1000.0),
            ("P1", 1100.0, 1050.0),
            ("P2", 1200.0, 1100.0),
            ("B", 1300.0, 1200.0),
        ]
        az_start = _azimuth(1000, 1000, 1100, 1050)
        az_end = _azimuth(1200, 1100, 1300, 1200)

        kwargs = dict(
            points=pts,
            start_azimuth=az_start,
            end_azimuth=az_end,
            grade=grade,
            num_angle_sets=2,
            seed=seed,
            target_closure_ratio=target_closure_ratio,
        )
        if start_ref and end_ref:
            kwargs["start_reference_point"] = start_ref
            kwargs["end_reference_point"] = end_ref

        return generate_traversing_workbook(**kwargs)

    def test_grade1_closed_adjustment(self):
        """一级闭合导线: 改正后终点坐标精确等于已知值."""
        wb = self._make_traverse(TraverseGrade.GRADE_1, seed=42,
                                 target_closure_ratio=0.3)
        comp = wb.computation
        assert comp is not None, "应有成果计算数据"

        # 验证改正后终点坐标精确归位
        last_point = comp.point_records[-1]
        assert last_point.corrected_x_m is not None
        assert last_point.corrected_y_m is not None
        # 终点已知坐标
        assert abs(last_point.corrected_x_m - 1300.0) < 1e-4, (
            f"改正后终点X={last_point.corrected_x_m}, 期望 1300.0"
        )
        assert abs(last_point.corrected_y_m - 1200.0) < 1e-4, (
            f"改正后终点Y={last_point.corrected_y_m}, 期望 1200.0"
        )

        # 验证 → 合规通过
        comp_report = check_traversing_compliance(wb)
        assert comp_report.passed, "合规检核未通过"

    def test_grade2_adjustment(self):
        """二级导线: 改正后终点坐标精确归位."""
        wb = self._make_traverse(TraverseGrade.GRADE_2, seed=42,
                                 target_closure_ratio=0.3)
        comp = wb.computation
        last_point = comp.point_records[-1]
        assert abs(last_point.corrected_x_m - 1300.0) < 1e-4
        assert abs(last_point.corrected_y_m - 1200.0) < 1e-4

    def test_root_adjustment(self):
        """图根导线: 改正后终点坐标精确归位."""
        wb = self._make_traverse(TraverseGrade.ROOT, seed=123,
                                 target_closure_ratio=0.3)
        comp = wb.computation
        last_point = comp.point_records[-1]
        assert abs(last_point.corrected_x_m - 1300.0) < 1e-4
        assert abs(last_point.corrected_y_m - 1200.0) < 1e-4

    def test_attached_traverse_adjustment(self):
        """附合导线 (含外部基准): 改正后终点坐标精确归位."""
        pts = [
            ("B", 1000.0, 1000.0),
            ("K1", 1100.0, 1050.0),
            ("K2", 1200.0, 1100.0),
            ("G", 1300.0, 1200.0),
        ]
        az_start = _azimuth(950, 980, 1000, 1000)  # B2 → B
        az_end = _azimuth(1300, 1200, 1350, 1230)   # G → G2
        start_ref = ("B2", 950.0, 980.0)
        end_ref = ("G2", 1350.0, 1230.0)

        wb = generate_traversing_workbook(
            points=pts,
            start_azimuth=az_start,
            end_azimuth=az_end,
            grade=TraverseGrade.GRADE_1,
            num_angle_sets=2,
            seed=42,
            target_closure_ratio=0.3,
            start_reference_point=start_ref,
            end_reference_point=end_ref,
        )
        comp = wb.computation
        last_point = comp.point_records[-1]
        assert abs(last_point.corrected_x_m - 1300.0) < 1e-4, (
            f"附合导线改正后终点X={last_point.corrected_x_m}, 期望 1300.0"
        )
        assert abs(last_point.corrected_y_m - 1200.0) < 1e-4, (
            f"附合导线改正后终点Y={last_point.corrected_y_m}, 期望 1200.0"
        )

    def test_math_true_mode_adjustment(self):
        """数学真值模式 (target_closure_ratio=0): 改正数极小但流程完整."""
        wb = self._make_traverse(TraverseGrade.GRADE_1, seed=42,
                                 target_closure_ratio=0.0)
        comp = wb.computation
        last_point = comp.point_records[-1]
        # 改正后坐标仍精确归位
        assert abs(last_point.corrected_x_m - 1300.0) < 1e-4
        assert abs(last_point.corrected_y_m - 1200.0) < 1e-4

        # 改正数极小 (数学真值模式下闭合差接近0)
        for rec in comp.point_records:
            if rec.angle_correction_rad is not None:
                # 角秒级改正数
                corr_arcsec = math.degrees(rec.angle_correction_rad) * 3600
                assert abs(corr_arcsec) < 5.0, (
                    f"数学真值模式角度改正数={corr_arcsec:.2f}s, 应极小"
                )

    def test_full_pipeline_with_output(self, tmp_path):
        """全流程: 生成 → 平差 → 合规 → 四种格式输出 (含成果表)."""
        wb = self._make_traverse(TraverseGrade.GRADE_1, seed=42,
                                 target_closure_ratio=0.3)

        # 验证通过
        val = validate_traversing_workbook(wb)
        assert val.all_passed

        # 合规通过
        comp = check_traversing_compliance(wb)
        assert comp.passed

        # JSON 含平差字段
        data = json.loads(workbook_to_json(wb))
        comp_data = data.get("computation", {})
        assert "point_records" in comp_data
        last_rec = comp_data["point_records"][-1]
        assert last_rec.get("corrected_x_m") is not None
        assert last_rec.get("corrected_y_m") is not None

        # Markdown 含成果计算表
        md = workbook_to_markdown(wb)
        assert "成果计算" in md

        # Excel 含成果计算 Sheet
        fp = str(tmp_path / "trav_adjusted.xlsx")
        workbook_to_excel(wb, fp)
        from openpyxl import load_workbook
        xlsx = load_workbook(fp)
        assert "成果计算" in xlsx.sheetnames


class TestLevelingAdjustmentE2E:
    """水准 E2E (含平差): 生成 → 平差 → 验证改正后高程闭合."""

    def test_grade3_single_run_adjustment(self):
        """三等单程: 改正后终点高程精确等于已知值."""
        route = RouteInfo("BM.A", 100.000, "BM.B", 108.000, 1.5)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=6, seed=42,
            target_closure_ratio=0.3,
        )
        assert wb.adjustment is not None
        records = wb.adjustment.records
        assert len(records) > 0

        # 改正后终点高程精确等于 108.0
        last_rec = records[-1]
        assert abs(last_rec.height_m - 108.0) < 1e-4, (
            f"改正后终点高程={last_rec.height_m}, 期望 108.0"
        )

        # 合规通过
        comp = check_leveling_compliance(wb)
        assert comp.passed

    def test_grade4_single_run_adjustment(self):
        """四等单程: 改正后终点高程精确归位."""
        route = RouteInfo("BM.C", 200.000, "BM.D", 215.000, 2.0)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_4,
            num_stations=8, seed=42,
            target_closure_ratio=0.3,
        )
        last_rec = wb.adjustment.records[-1]
        assert abs(last_rec.height_m - 215.0) < 1e-4

    def test_grade2_round_trip_adjustment(self):
        """二等往返测: 改正后终点高程精确归位."""
        route = RouteInfo("BM.E", 50.000, "BM.F", 50.800, 0.5)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_2,
            num_stations=6, seed=42,
            round_trip=True,
            target_round_trip_ratio=0.4,
        )
        assert wb.adjustment is not None
        last_rec = wb.adjustment.records[-1]
        assert abs(last_rec.height_m - 50.8) < 1e-4, (
            f"往返测改正后终点高程={last_rec.height_m}, 期望 50.8"
        )

        # 往返测附注存在
        assert wb.adjustment.round_trip_discrepancy_mm is not None
        assert wb.adjustment.mean_height_diff_m is not None

    def test_math_true_mode_adjustment(self):
        """数学真值模式: 改正数极小但流程完整."""
        route = RouteInfo("BM.A", 100.000, "BM.B", 108.000, 1.5)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=6, seed=42,
            target_closure_ratio=0.0,
        )
        last_rec = wb.adjustment.records[-1]
        assert abs(last_rec.height_m - 108.0) < 1e-4

        # 改正数极小
        for rec in wb.adjustment.records:
            if rec.correction_mm is not None:
                assert abs(rec.correction_mm) < 1.0, (
                    f"数学真值模式改正数={rec.correction_mm:.3f}mm, 应极小"
                )

    def test_full_pipeline_with_output(self, tmp_path):
        """全流程: 生成 → 平差 → 合规 → 四种格式输出 (含成果表)."""
        route = RouteInfo("BM.A", 100.000, "BM.B", 108.000, 1.5)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=6, seed=42,
            target_closure_ratio=0.3,
        )

        # 验证通过
        val = validate_leveling_workbook(wb)
        assert val.all_passed

        # 合规通过
        comp = check_leveling_compliance(wb)
        assert comp.passed

        # JSON 含平差字段
        data = json.loads(workbook_to_json(wb))
        assert "adjustment" in data
        adj = data["adjustment"]
        assert adj["records"][-1]["height_m"] is not None

        # Markdown 含成果计算表
        md = workbook_to_markdown(wb)
        assert "成果计算" in md

        # Excel 含成果计算 Sheet
        fp = str(tmp_path / "lev_adjusted.xlsx")
        workbook_to_excel(wb, fp)
        from openpyxl import load_workbook
        xlsx = load_workbook(fp)
        assert "成果计算" in xlsx.sheetnames


class TestControlledClosureE2E:
    """可控非零闭合差模式 E2E: 生成 → 平差 → 检核全流程."""

    def test_traversing_controlled_closure(self):
        """导线可控闭合差: 闭合差非零但合规, 平差后精确归位."""
        pts = [
            ("A", 1000.0, 1000.0),
            ("P1", 1100.0, 1050.0),
            ("P2", 1200.0, 1100.0),
            ("B", 1300.0, 1200.0),
        ]
        az_start = _azimuth(1000, 1000, 1100, 1050)
        az_end = _azimuth(1200, 1100, 1300, 1200)

        wb = generate_traversing_workbook(
            points=pts, start_azimuth=az_start, end_azimuth=az_end,
            grade=TraverseGrade.GRADE_1, seed=42,
            target_closure_ratio=0.5,  # 50% 限差
        )
        comp = wb.computation

        # 闭合差非零 (可控模式)
        assert abs(comp.fd_m) > 1e-6, "可控模式下闭合差应非零"

        # 合规通过 (在限差内)
        comp_report = check_traversing_compliance(wb)
        assert comp_report.passed, "可控闭合差应合规"

        # 平差后精确归位
        last_point = comp.point_records[-1]
        assert abs(last_point.corrected_x_m - 1300.0) < 1e-4
        assert abs(last_point.corrected_y_m - 1200.0) < 1e-4

    def test_leveling_controlled_closure(self):
        """水准可控闭合差: 闭合差非零但合规, 平差后精确归位."""
        route = RouteInfo("BM.A", 100.000, "BM.B", 108.000, 1.5)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_3,
            num_stations=6, seed=42,
            target_closure_ratio=0.3,
        )

        # 闭合差非零
        assert wb.adjustment.closure_error_mm is not None
        assert abs(wb.adjustment.closure_error_mm) > 0.01, "可控模式下闭合差应非零"

        # 合规通过
        comp = check_leveling_compliance(wb)
        assert comp.passed

        # 平差后精确归位
        last_rec = wb.adjustment.records[-1]
        assert abs(last_rec.height_m - 108.0) < 1e-4

    def test_round_trip_controlled_closure(self):
        """往返测可控不符值: 不符值非零但合规, 平差后精确归位."""
        route = RouteInfo("BM.A", 100.000, "BM.B", 108.000, 1.5)
        wb = generate_leveling_workbook(
            route=route, grade=LevelingGrade.GRADE_2,
            num_stations=8, seed=42,
            round_trip=True,
            target_round_trip_ratio=0.5,
        )

        # 往返测不符值非零
        assert wb.round_trip_discrepancy_mm is not None
        assert abs(wb.round_trip_discrepancy_mm) > 0.01, "可控模式下不符值应非零"

        # 合规通过
        comp = check_leveling_compliance(wb)
        assert comp.passed

        # 平差后精确归位
        last_rec = wb.adjustment.records[-1]
        assert abs(last_rec.height_m - 108.0) < 1e-4
